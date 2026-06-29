"""
trip_xlsx.py — reusable workbook builder for the trip-planner skill (Phase 5: Assembly).

Why this exists
---------------
Phase 5 turns the researched data (hotels, flights, activities, transport, dates)
into a polished, formula-driven Excel workbook. The *styling and plumbing* of that
workbook — fonts, fills, borders, currency formats, hyperlink mechanics, the link
audit, and the recalc step — are the same on every trip and are fiddly to get right
from scratch (the classic bug: coloring a cell blue but forgetting to set
`cell.hyperlink`, so it isn't actually clickable). Rewriting ~400 lines of that
boilerplate per run is slow and error-prone, so it lives here once.

What this does NOT do
---------------------
It does not decide which tabs a trip needs or how the cost formulas read — that's
trip-specific and stays in the orchestrator's build script. A ski trip wants a
ski-pass tab; a beach trip won't. This module gives you primitives and a couple of
table helpers; you compose the tabs.

Typical usage (orchestrator writes a short script like this)
------------------------------------------------------------
    import sys
    sys.path.insert(0, "<skill-dir>/scripts")
    import trip_xlsx as tx

    wb = tx.new_workbook()
    ws = tx.sheet(wb, "Hotels", first=True,
                  widths={1: 26, 2: 12, 3: 10, 4: 12, 5: 12})

    tx.title(ws, 1, "HOTELS — Paris", ncols=5)
    tx.header(ws, 3, ["Property", "Platform", "Score", "Total stay", "Total (home)"])
    r = 4
    tx.link(ws, r, 1, "Hotel Lutetia", "https://www.booking.com/hotel/...")
    tx.lbl(ws, r, 2, "Booking.com")
    tx.lbl(ws, r, 3, "9.1", align="center")
    tx.num(ws, r, 4, 1840, fmt=tx.money("EUR"))
    tx.formula(ws, r, 5, f"=D{r}*'Trip Summary'!$C$11")
    tx.zebra(ws, 4, r, 1, 5)
    tx.borders(ws, 3, 1, r, 5)

    # Save + audit links + recalc, all in one call:
    report = tx.finalize(wb, "/mnt/user-data/outputs/Paris_Trip.xlsx",
                         link_checks=[("Hotels", "D")])   # (sheet, price column)
    print(report)

Every cell helper returns the cell, so you can tweak it further if needed.
"""

from __future__ import annotations

import json
import os
import re
import subprocess
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette ──────────────────────────────────────────────────────────────────
# Override any of these after import if a trip wants a different look:
#   import trip_xlsx as tx; tx.DARK = "1F4E79"
DARK = "14505C"        # title band fill
MID = "2E7D8A"         # section / header band fill
LIGHT = "EAF3F5"       # zebra stripe (alternate rows)
ACCENT = "C9A227"      # emphasis (e.g. per-person row)
WHITE = "FFFFFF"
CHOSEN = "FBF3D5"      # highlight fill for the recommended/chosen row
NOTE_GRAY = "F2F2F2"   # subtle fill for note/estimate rows
# The glyph the user puts in a "Choose" column to pick which option (hotel,
# flight, …) feeds the Trip Summary total. Driven by a dropdown + SUMIF lookup
# (see choose_validation / chosen_total below), so switching the pick is a
# one-click change that recalculates the total — no formula editing.
CHOSEN_MARK = "✓"

# ── Font colors (text, not fills) ────────────────────────────────────────────
C_TEXT = "000000"      # default text / formulas
C_INPUT = "0000FF"     # hardcoded numbers the user may edit (blue)
C_XREF = "375623"      # cross-sheet reference formulas (green)
C_LINK = "1155CC"      # hyperlinks (blue, underlined)
C_NOTE = "808080"      # gray italic notes / estimates
C_WHITE = "FFFFFF"
C_GOOD = "375623"      # affirmative inline text (e.g. "included")
C_WARN = "C00000"      # cautionary inline text (e.g. "extra fee")

# ── Number formats ───────────────────────────────────────────────────────────
FMT_INT = "0"
FMT_TEXT = "@"

# Currency symbols for the formats commonly needed. Extend as required.
_CURRENCY_SYMBOL = {
    "CAD": "C$", "USD": "$", "EUR": "€", "GBP": "£", "AUD": "A$",
    "NZD": "NZ$", "JPY": "¥", "CHF": "CHF ", "MXN": "$", "INR": "₹",
}


def money(currency: str = "USD") -> str:
    """Return an Excel number format for a currency code, e.g. money("EUR")."""
    sym = _CURRENCY_SYMBOL.get((currency or "").upper(), (currency or "") + " ")
    # positive ; negative-in-parens ; zero-as-dash
    return f'"{sym}"#,##0_);("{sym}"#,##0);-'


# Backwards-friendly default used widely in build scripts:
FMT_MONEY = money("USD")

_THIN = Side(style="thin", color="D0D9DB")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ── Core cell writer ─────────────────────────────────────────────────────────
def cell(ws, r, c, value=None, *, bold=False, size=10, color=C_TEXT, fill=None,
         align="left", wrap=False, fmt=None, italic=False, underline=None,
         border=True):
    """Write a value to (r, c) with Arial styling. Returns the cell.

    This is the single styling primitive; the named helpers below are thin
    wrappers around it for readability.
    """
    cl = ws.cell(row=r, column=c, value=value)
    cl.font = Font(name="Arial", bold=bold, size=size, color=color,
                   italic=italic, underline=underline)
    cl.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cl.fill = PatternFill("solid", start_color=fill)
    if fmt:
        cl.number_format = fmt
    if border:
        cl.border = BORDER
    return cl


# ── Named cell helpers ───────────────────────────────────────────────────────
def lbl(ws, r, c, text, *, bold=False, align="left", fill=None, wrap=False,
        color=C_TEXT, size=10):
    """Plain text label."""
    return cell(ws, r, c, text, bold=bold, align=align, fill=fill, wrap=wrap,
                color=color, size=size, fmt=FMT_TEXT if isinstance(text, str) else None)


def num(ws, r, c, value, *, fmt=FMT_MONEY, color=C_INPUT, fill=None, align="right"):
    """Hardcoded number the user may edit — blue by default."""
    return cell(ws, r, c, value, color=color, fill=fill, align=align, fmt=fmt)


def formula(ws, r, c, expr, *, fmt=FMT_MONEY, color=C_TEXT, fill=None, align="right"):
    """A computed formula — black by default. `expr` must start with '='."""
    if not (isinstance(expr, str) and expr.startswith("=")):
        raise ValueError(f"formula() expects a string starting with '=', got: {expr!r}")
    return cell(ws, r, c, expr, color=color, fill=fill, align=align, fmt=fmt)


def xref(ws, r, c, expr, *, fmt=FMT_MONEY, fill=None, align="right"):
    """A cross-sheet reference formula — green, so the user can see it's a link
    to a chosen value elsewhere (e.g. ='Hotels'!J7)."""
    return formula(ws, r, c, expr, fmt=fmt, color=C_XREF, fill=fill, align=align)


def link(ws, r, c, text, url, *, bold=False, fill=None, wrap=True, align="left"):
    """A real, clickable hyperlink. Sets BOTH cell.hyperlink and link styling —
    coloring text blue alone does not make it clickable.

    If `url` is falsy this raises, because a priced row must never ship without a
    working link (see the skill's link-coverage gate). Pass a search URL for the
    exact item rather than leaving it blank.
    """
    if not url:
        raise ValueError(
            f"link() called with empty url for {text!r} at {get_column_letter(c)}{r}. "
            "Every booking/source cell needs a real URL — use a search URL for the "
            "exact item if no deep link exists."
        )
    cl = cell(ws, r, c, text, bold=bold, color=C_LINK, underline="single",
              fill=fill, wrap=wrap, align=align, fmt=FMT_TEXT)
    cl.hyperlink = url
    return cl


def note(ws, r, c, text, *, colspan=1, fill=None):
    """Gray italic note / source / estimate flag. Optionally merges colspan cells."""
    cl = cell(ws, r, c, text, color=C_NOTE, size=9, italic=True, wrap=True,
              align="left", fill=fill, border=False)
    cl.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if colspan > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + colspan - 1)
    return cl


# ── Layout helpers ───────────────────────────────────────────────────────────
_INVALID_SHEET = re.compile(r'[:\\/?*\[\]]')


def _safe_title(title: str) -> str:
    """Excel sheet titles: <=31 chars, no : \\ / ? * [ ]. Emoji are allowed."""
    t = _INVALID_SHEET.sub(" ", title).strip()
    return t[:31] if len(t) > 31 else t


def new_workbook() -> Workbook:
    """Fresh workbook. Use sheet(..., first=True) for the first tab so you don't
    leave a stray empty 'Sheet' behind."""
    return Workbook()


def sheet(wb, title, *, widths=None, first=False, hide_gridlines=True):
    """Create (or, with first=True, repurpose the initial blank sheet) a tab.
    `widths` is an optional {col_index_or_letter: width} dict."""
    title = _safe_title(title)
    if first:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title)
    if hide_gridlines:
        ws.sheet_view.showGridLines = False
    if widths:
        set_col_widths(ws, widths)
    return ws


def set_col_widths(ws, widths):
    """widths: {1: 26, 'B': 12, ...} — keys may be 1-based indexes or letters."""
    for key, w in widths.items():
        col = key if isinstance(key, str) else get_column_letter(key)
        ws.column_dimensions[col].width = w


def _band(ws, r, c, text, ncols, *, fill, fg, size, height):
    cl = cell(ws, r, c, text, bold=True, color=fg, fill=fill, size=size,
              align="center", wrap=True)
    if ncols > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + ncols - 1)
        # carry the fill across the merged span so the whole band is colored
        for cc in range(c + 1, c + ncols):
            ws.cell(r, cc).fill = PatternFill("solid", start_color=fill)
    if height:
        ws.row_dimensions[r].height = height
    return cl


def title(ws, r, text, *, ncols=1, start_col=1, fill=DARK, fg=C_WHITE, size=13,
          height=28):
    """Big title band across `ncols` columns."""
    return _band(ws, r, start_col, text, ncols, fill=fill, fg=fg, size=size, height=height)


def band(ws, r, text, *, ncols=1, start_col=1, fill=MID, fg=C_WHITE, size=10,
         height=22):
    """Section band (sub-header)."""
    return _band(ws, r, start_col, text, ncols, fill=fill, fg=fg, size=size, height=height)


def header(ws, r, headers, *, start_col=1, fill=MID, fg=C_WHITE, height=26):
    """Write a row of column headers (bold, centered, wrapped, banded)."""
    for i, h in enumerate(headers):
        cell(ws, r, start_col + i, h, bold=True, color=fg, fill=fill,
             align="center", wrap=True)
    if height:
        ws.row_dimensions[r].height = height
    return r


def fill_row(ws, r, c1, c2, color):
    f = PatternFill("solid", start_color=color)
    for c in range(c1, c2 + 1):
        ws.cell(r, c).fill = f


def zebra(ws, r1, r2, c1, c2, colors=(LIGHT, WHITE)):
    """Apply alternating row fills across [r1, r2] x [c1, c2].
    Skips any cell that already has a non-default fill (so a CHOSEN highlight or a
    manual fill_row is preserved)."""
    for i, r in enumerate(range(r1, r2 + 1)):
        color = colors[i % len(colors)]
        for c in range(c1, c2 + 1):
            cl = ws.cell(r, c)
            existing = cl.fill
            already = (existing is not None and existing.patternType == "solid"
                       and (existing.fgColor.rgb not in (None, "00000000")))
            if not already:
                cl.fill = PatternFill("solid", start_color=color)


def borders(ws, r1, c1, r2, c2):
    for row in ws.iter_rows(min_row=r1, min_col=c1, max_row=r2, max_col=c2):
        for cl in row:
            cl.border = BORDER


def highlight_row(ws, r, c1, c2, color=CHOSEN):
    """Mark the chosen/recommended option row."""
    fill_row(ws, r, c1, c2, color)

# ── Choose-an-option mechanism (marker column + lookup) ──────────────────────
# A segment (a city's hotels, the flights for a leg, …) lists several options,
# one per row. A "Choose" column lets the user mark exactly one row; the Trip
# Summary pulls *that* row's total via SUMIF, so picking a different option is a
# one-click change that recalculates the whole trip — no formula editing.
#
# Build it in three steps:
#   1. choose_validation(ws, r1, r2, c)         → dropdown on the marker cells
#   2. put CHOSEN_MARK in the recommended row's marker cell (pre-select it) and
#      highlight_row() that row so the recommended pick is visible by default
#   3. on Trip Summary: chosen_total(value_rng, marker_rng, sheet="Hotels")
#      → a SUMIF that totals whichever row is marked
def choose_validation(ws, r1, r2, c, *, marker=CHOSEN_MARK):
    """Add a one-option dropdown (marker, or blank) to the Choose-column cells in
    rows r1..r2 of column c. The dropdown keeps the user to a valid marker and
    makes the column obviously interactive. Returns the DataValidation."""
    col = get_column_letter(c) if isinstance(c, int) else c
    dv = DataValidation(type="list", formula1=f'"{marker}"', allow_blank=True)
    dv.prompt = "Pick one option for this segment — it feeds the Trip Summary total."
    dv.promptTitle = "Choose this option"
    dv.add(f"{col}{r1}:{col}{r2}")
    ws.add_data_validation(dv)
    return dv


def choose_mark(ws, r, c, *, marker=CHOSEN_MARK):
    """Write the chosen marker into a Choose cell (centered). Use on the
    recommended row so the workbook opens with a valid default selection."""
    return cell(ws, r, c, marker, align="center", bold=True, color=C_XREF)


def _rng(value_range, sheet=None):
    return f"'{sheet}'!{value_range}" if sheet else value_range


def chosen_total(value_range, marker_range, *, sheet=None, marker=CHOSEN_MARK):
    """Return a SUMIF formula that totals the value of whichever row is marked in
    the Choose column. Only one row should be marked, so the SUM is that row's
    value; switching the mark re-points the total automatically.

        # On Trip Summary, pull the chosen Paris hotel's home-currency total:
        tx.formula(ws, r, c,
            tx.chosen_total("$I$4:$I$8", "$J$4:$J$8", sheet="Hotels"),
            color=tx.C_XREF)

    `value_range` / `marker_range` are A1 ranges (use $ to keep them absolute).
    Pass `sheet` when the lookup lives on a different tab than the data."""
    vr = _rng(value_range, sheet)
    mr = _rng(marker_range, sheet)
    return f'=SUMIF({mr},"{marker}",{vr})'


def chosen_label(value_range, marker_range, *, sheet=None, marker=CHOSEN_MARK):
    """Like chosen_total but returns the *text* (e.g. the chosen property/airline
    name) of the marked row, via INDEX/MATCH. Handy for echoing the current pick
    into the Trip Summary's Details column so it updates with the marker. (Uses
    INDEX/MATCH, not XLOOKUP, so it recalculates in Excel and LibreOffice alike.)"""
    vr = _rng(value_range, sheet)
    mr = _rng(marker_range, sheet)
    return f'=INDEX({vr},MATCH("{marker}",{mr},0))'


# ── Per-person figures ───────────────────────────────────────────────────────
def per_person(total_cell, travelers_cell):
    """Return a formula dividing a group total by the travelers count, e.g.
    per_person("F12", "$C$4") → '=F12/$C$4'. Use to add a per-person column that
    sits *before* the group total so a single flight/room/seat price reads at a
    glance. travelers_cell should be absolute ($) so it survives fill-down."""
    return f"={total_cell}/{travelers_cell}"


# ── Link-coverage audit ──────────────────────────────────────────────────────
def audit_links(wb_or_path, checks):
    """Verify every PRICED row on the given sheets carries a clickable hyperlink.

    checks: list of (sheet_name, price_col) where price_col is the column letter
            (or 1-based index) whose presence of a number or formula marks a row
            as "priced". A row counts as covered if ANY cell in that row has a
            hyperlink, so it doesn't matter whether the link sits on the name
            column or a dedicated "Book" column.

    Pass only the bookable tabs (Hotels, Flights, Activities, Transport...) — not
    the Trip Summary, whose per-person formula cells are priced but shouldn't link.

    Returns a list of human-readable strings for offending rows (empty == all good).
    """
    wb = load_workbook(wb_or_path) if isinstance(wb_or_path, (str, os.PathLike)) else wb_or_path
    missing = []
    for sheet_name, price_col in checks:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        pc = price_col if isinstance(price_col, int) else column_index_from_string(price_col)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            if len(row) < pc:
                continue
            price = row[pc - 1].value
            is_number = isinstance(price, (int, float))
            is_formula = isinstance(price, str) and price.startswith("=")
            if not (is_number or is_formula):
                continue  # not a priced row
            if not any(c.hyperlink for c in row):
                missing.append(f"{sheet_name}!row {row[0].row} (priced, no clickable link)")
    return missing


# ── Live link sanity check ───────────────────────────────────────────────────
# audit_links() only proves a link EXISTS. check_links() proves each one actually
# RESOLVES: it fetches every hyperlink and classifies the result. The wrinkle is
# that the big booking sites (Booking.com, Expedia, Tripadvisor, lastminute.com,
# and friends) routinely answer automated requests with 403/429/503 or a timeout
# even though the URL is perfectly good in a real browser — so a naive "anything
# that isn't 200 is broken" check would cry wolf. We therefore split the verdict
# three ways:
#   • ok      — reachable, and the page does not read as a dead / "not found" page
#   • blocked — refused by a known bot-blocking travel domain (nothing to fix; the
#               link is fine for the human who clicks it)
#   • broken  — genuinely wrong: DNS / connection failure, 404 / 410, a 5xx, or a
#               200 whose CONTENT says the page is gone ("page not found", etc.)
# Only "broken" links need a fix-task; "blocked" links are reported but accepted.

BOT_BLOCKING_DOMAINS = (
    "booking.com", "expedia.", "tripadvisor.", "lastminute.com", "hotels.com",
    "airbnb.", "agoda.", "kayak.", "vrbo.", "skyscanner.", "getyourguide.",
    "viator.", "google.com/travel",
)

# Strong "this page is dead" phrases. Matched against the <title> first (error
# pages almost always say it there) and then the first few KB of the body, so a
# 200-with-an-error-page (a "soft 404") is still caught.
SOFT_404_MARKERS = (
    "page not found", "404 not found", "page you requested could not be found",
    "page you were looking for", "no longer available", "is not available",
    "this page doesn't exist", "this page does not exist", "page cannot be found",
    "page can't be found", "we couldn't find", "could not be found",
    "sorry, we can't find", "no results found",
)


def _host_blocks_bots(url):
    u = (url or "").lower()
    return any(d in u for d in BOT_BLOCKING_DOMAINS)


def _soft_404_reason(body):
    """Return the dead-page phrase found in the page, or None. Prefers the
    <title>; falls back to the first 4 KB of the body."""
    if not body:
        return None
    low = body.lower()
    m = re.search(r"<title[^>]*>(.*?)</title>", low, re.S)
    title = m.group(1) if m else ""
    for phrase in SOFT_404_MARKERS:
        if phrase in title:
            return phrase
    head = low[:4000]
    for phrase in SOFT_404_MARKERS:
        if phrase in head:
            return phrase
    return None


def classify_link(url, status, body, error=None):
    """Pure classifier (no network) → (verdict, reason), verdict ∈
    {"ok","blocked","broken"}. It's deliberately side-effect-free so it can be
    unit tested offline; check_links() feeds it live (status, body, error)."""
    blocks = _host_blocks_bots(url)
    if error is not None:
        msg = str(error).lower()
        timed_out = "timed out" in msg or "timeout" in msg
        if blocks and (timed_out or "forbidden" in msg or "403" in msg):
            return "blocked", f"{error} — bot-blocking travel domain, fine in a browser"
        return "broken", f"could not connect: {error}"
    if status in (403, 429, 503):
        if blocks:
            return "blocked", f"HTTP {status} from a bot-blocking travel domain (valid in a browser)"
        return "broken", f"HTTP {status} (access denied / rate-limited)"
    if status in (404, 410):
        return "broken", f"HTTP {status} (page gone)"
    if status is not None and 500 <= status < 600:
        return "broken", f"HTTP {status} (server error)"
    if status is not None and 200 <= status < 400:
        reason = _soft_404_reason(body)
        if reason:
            return "broken", f'HTTP {status} but the page reads as dead ("{reason}")'
        return "ok", f"HTTP {status}"
    return "broken", f"unexpected response (status={status})"


def _default_fetch(url, timeout=12):
    """Fetch a URL like a browser would; return (status, body, error). Network
    errors (DNS, refused, timeout) come back as the error slot, not an exception."""
    import urllib.request
    import urllib.error
    req = urllib.request.Request(url, method="GET", headers={
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            status = getattr(resp, "status", None) or resp.getcode()
            return status, resp.read(20000).decode("utf-8", "replace"), None
    except urllib.error.HTTPError as e:
        body = ""
        try:
            body = e.read(20000).decode("utf-8", "replace")
        except Exception:
            pass
        return e.code, body, None
    except Exception as e:  # URLError (DNS / refused), socket timeout, etc.
        return None, "", e


def collect_links(wb_or_path):
    """Map every hyperlink in the workbook to where it lives:
    {url: [{"sheet","cell","text"}, ...]}. URLs are deduplicated, so a link that
    appears on several rows is only fetched once."""
    wb = load_workbook(wb_or_path) if isinstance(wb_or_path, (str, os.PathLike)) else wb_or_path
    links = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.hyperlink and c.hyperlink.target:
                    links.setdefault(c.hyperlink.target, []).append(
                        {"sheet": ws.title, "cell": c.coordinate, "text": c.value})
    return links


def check_links(wb_or_path, *, fetcher=None, max_workers=8):
    """Fetch and classify every hyperlink in the workbook. Returns a report:

        {"checked": int,
         "ok":      [ {url, locations} ... ],
         "blocked": [ {url, reason, locations} ... ],   # accepted, no fix needed
         "broken":  [ {url, reason, locations} ... ]}   # each needs a fix-task

    `fetcher(url) -> (status, body, error)` is injectable so the classification
    can be unit tested offline; the default hits the network with a browser
    User-Agent. Drive fix-tasks off the "broken" list only — "blocked" is
    informational (a bot-blocking site, fine for a human in a browser)."""
    fetcher = fetcher or _default_fetch
    links = collect_links(wb_or_path)
    report = {"checked": len(links), "ok": [], "blocked": [], "broken": []}

    def _one(url):
        status, body, error = fetcher(url)
        verdict, reason = classify_link(url, status, body, error)
        return url, verdict, reason

    urls = list(links.keys())
    if max_workers and max_workers > 1 and len(urls) > 1:
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            results = list(ex.map(_one, urls))
    else:
        results = [_one(u) for u in urls]

    for url, verdict, reason in results:
        entry = {"url": url, "locations": links[url]}
        if verdict != "ok":
            entry["reason"] = reason
        report[verdict].append(entry)
    return report


# ── Recalc + finalize ────────────────────────────────────────────────────────
def _find_recalc():
    for p in (
        "/mnt/skills/public/xlsx/scripts/recalc.py",
        "/mnt/skills/private/xlsx/scripts/recalc.py",
    ):
        if os.path.exists(p):
            return p
    return None


def recalc(path):
    """Run the xlsx skill's recalc.py on the file if available.
    Returns (ok: bool, output: str). ok is None if recalc.py wasn't found."""
    script = _find_recalc()
    if not script:
        return None, "recalc.py not found — run the xlsx skill's recalc manually."
    proc = subprocess.run([sys.executable, script, path],
                          capture_output=True, text=True)
    out = ((proc.stdout or "") + (proc.stderr or "")).strip()
    # recalc.py prints JSON like {"status": "success", "total_errors": 0, ...}.
    # Parse it rather than substring-matching "error" (which appears in the keys).
    ok = proc.returncode == 0
    try:
        data = json.loads(out)
        ok = (data.get("status") == "success") and (data.get("total_errors", 0) == 0)
    except (json.JSONDecodeError, AttributeError):
        # Couldn't parse — fall back to: success only if no obvious failure text.
        ok = proc.returncode == 0 and "traceback" not in out.lower()
    return ok, out


def finalize(wb, out_path, *, link_checks=None, run_recalc=True):
    """Save the workbook, audit links, and run recalc — the standard close-out.

    out_path:    where to write (parent dirs are created if missing).
    link_checks: list of (sheet, price_col) for audit_links; pass the bookable
                 tabs. If None, the audit is skipped (do it manually).
    run_recalc:  run the xlsx skill's recalc.py and report formula errors.

    Returns a report dict. Inspect it: if report["missing_links"] is non-empty or
    report["recalc_ok"] is False, fix and call finalize() again before presenting.
    Never present a workbook with missing links or formula errors.
    """
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)

    report = {"path": out_path, "missing_links": [], "recalc_ok": None, "recalc_output": ""}

    if link_checks:
        report["missing_links"] = audit_links(out_path, link_checks)

    if run_recalc:
        ok, out = recalc(out_path)
        report["recalc_ok"] = ok
        report["recalc_output"] = out

    return report
